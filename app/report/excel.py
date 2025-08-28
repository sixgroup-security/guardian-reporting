# This file is part of Guardian.
#
# Guardian is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Guardian is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Guardian. If not, see <https://www.gnu.org/licenses/>.

import os
import re
import logging
from copy import copy
from core.config import Settings
from typing import Callable, Dict, Tuple
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from sqlalchemy.orm import Session
from .util import ReportCreatorBase
from schema import SessionLocal
from schema.project import ProjectReport
from schema.application import Application
from schema.reporting.report_template import ReportTemplateFileVersion
from schema.tagging.mitre_cwe import CweBaseRelationship, CweCategory, CweWeakness
from schema.reporting.report_section_management.vulnerability import VulnerabilityStatus

__author__ = "Lukas Reiter"
__copyright__ = "Copyright (C) 2024 Lukas Reiter"
__license__ = "GPLv3"

logger = logging.getLogger(__name__)


class ReportCreator(ReportCreatorBase):
    """
    This class is responsible for creating the Excel report.
    """
    COLUMN_NAMES = [
        "Application ID",
        "Application Name",
        "ID",
        "Title",
        "Status",
        "Description",
        "Measure Title",
        "Measure",
        "Severity",
        "CVSS Score",
        "CVSS Vector"
    ]

    def __init__(
            self,
            excel_file: str,
            **kwargs
    ):
        super().__init__(**kwargs)
        self.excel_file = excel_file
        self._re_cvs_injection = re.compile(r"^([^a-zA-Z0-9])", re.MULTILINE)

    def test_cvs_injection(self, markdown: str) -> str:
        """
        This method checks for unauthorized LaTeX commands in the given Markdown text.
        """
        def cvs_injection(match):
            result = match.group(0)
            cvs_command = match.group(1)
            if cvs_command:
                self._logger.warning(
                    f"Markdown starts with unauthorized special character '{result}' in report '{self.report.id}' "
                    f"and version '{self.latest_version_info.version}'"
                )
                result = ""
            return result
        return self._re_cvs_injection.sub(cvs_injection, markdown)

    def get_text(
            self,
            markdown: str | None,
            pre_placeholder_fn: Callable[[str, Dict[str, str], str, str | None], str | None] | None = None,
    ):
        """
        Prepares the content for Excel.
        """
        placeholder_fn = pre_placeholder_fn if pre_placeholder_fn else self.replace_placeholders_only_func
        result = (markdown or "").strip()
        # We need to replace placeholders with the final values.
        result = self.replace_placeholders(
            report_text=result,
            placeholder_values=self.placeholders,
            placeholder_pattern=self.pre_placeholder_pattern,
            placeholder_fn=placeholder_fn
        )
        result = self.test_cvs_injection(result)
        return result

    def get_xlsx(self) -> bytes:
        """
        Returns the Excel file as bytes.
        """
        with open(self.excel_file, "rb") as file:
            return file.read()

    @staticmethod
    def get_application_info(session: Session, project: ProjectReport) -> Tuple[str, str]:
        """
        Normalizes application, component information.
        """
        application_name = ""
        component_name = ""
        if len(project.applications) > 1:
            return "", ""
        if application := project.applications[0] if project.applications else None:
            if application.type in [ApplicationType.application_component, ApplicationType.platform_component]:
                component_name = application.name
                # TODO: This information should actually be optained from the JSON object and not the database
                if parent := session.query(Application).filter_by(id=application.parent_application_id).first():
                    application_name = parent.name
            else:
                application_name = application.name
        return application_name, component_name

    # Function to copy cell styles
    def copy_cell_style(self, src_cell, tgt_cell, ws):
        if src_cell.has_style:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = copy(src_cell.number_format)
            tgt_cell.protection = copy(src_cell.protection)
            tgt_cell.alignment = copy(src_cell.alignment)
        # Copy data validation
        for dv in ws.data_validations.dataValidation:
            if f"{src_cell.column_letter}{self.settings.excel_template_row}" in dv.cells:
                dv.add(tgt_cell)

    def create(self):
        """
        Creates the Excel file based on the given template file.
        """
        template_file = self.settings.get_excel_template_file(self.info.project.report.version)
        latest_version = self.latest_version_info
        # The BU contains the application owner
        business_unit = self.info.project.application_owner.abbreviation or ""
        with SessionLocal() as session:
            application_name, component_name = self.get_application_info(session=session, project=self.info.project)
            sheet_name = self.settings.excel_sheet_name
            if not os.path.isfile(template_file):
                raise FileNotFoundError(f"Excel template file '{template_file}' does not exist.")
            workbook = load_workbook(template_file)
            ws = workbook[sheet_name]
            tb = ws.tables[self.settings.excel_table_name]
            from_x, from_y, to_x, to_y = range_boundaries(tb.ref)
            row = from_y + 1
            # Populate cells
            for section in self.info.project.report.sections:
                for vulnerability in section.vulnerabilities:
                    if not vulnerability.visible or vulnerability.status in [VulnerabilityStatus.resolved]:
                        continue
                    cwe_category = ""
                    if vulnerability.cwe_weakness:
                        if (category := session.query(CweCategory)
                                .join(CweBaseRelationship, CweCategory.id == CweBaseRelationship.destination_id)
                                .join(CweWeakness, CweWeakness.id == CweBaseRelationship.source_id)
                                .filter(CweWeakness.id == vulnerability.cwe_weakness.id)
                                .order_by(CweCategory.cwe_id).first()):
                            cwe_category = category.name
                    # We need to be consistent with the PDF creation
                    if vulnerability.cvss_score and vulnerability.cvss_score > 0 and vulnerability.cvss_vector:
                        cvss_vector = vulnerability.cvss_vector
                        cvss_score = vulnerability.cvss_score
                    else:
                        cvss_vector = ""
                        cvss_score = ""
                    ws[f'A{row}'] = ", ".join(
                        [self.get_text(item.application_id) for item in self.info.project.applications]
                    )  # Application ID
                    ws[f'B{row}'] = ", ".join(
                        [self.get_text(item.name) for item in self.info.project.applications]
                    )  # Application Name
                    ws[f'C{row}'] = vulnerability.vulnerability_id_str  # D
                    ws[f'D{row}'] = self.get_text(vulnerability.name)  # Title
                    ws[f'E{row}'] = vulnerability.status_str  # Status
                    ws[f'F{row}'] = self.get_text(vulnerability.description)  # Description
                    ws[f'G{row}'] = self.get_text(vulnerability.measure_title)  # Measure Title
                    ws[f'H{row}'] = self.get_text(vulnerability.measure_recommendation)  # Measure
                    ws[f'I{row}'] = vulnerability.severity_str  # Severity
                    ws[f'J{row}'] = vulnerability.cvss_score  # CVSS Score
                    ws[f'K{row}'] = self.get_text(vulnerability.cvss_vector)  # CVSS Vector
                    row += 1
                    # Apply formatting for all cells in current row
                    if row != (from_y + 1):
                        for x in range(0, len(self.COLUMN_NAMES)):
                            template_cell = ws.cell(row=from_y + self.settings.excel_template_row, column=from_x + x)
                            new_cell = ws.cell(row=row, column=from_x + x)
                            self.copy_cell_style(template_cell, new_cell, ws)
                    row += 1
            # Update table definition
            tb.ref = f"{get_column_letter(from_x)}{from_y}:{get_column_letter(to_x)}{row - 1}"
            # Save the file
            workbook.save(self.excel_file)

    @staticmethod
    def check(settings: Settings):
        """
        Checks prerequisites for creating Excel files.
        """
        # Check if Excel template file exist
        for version in ReportTemplateFileVersion:
            file_name = settings.get_excel_template_file(version)
            if not os.path.isfile(file_name):
                raise FileNotFoundError(f"Excel template file '{file_name}' not found.")
            workbook = load_workbook(file_name)
            # Check if Sheet exist
            if settings.excel_sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet '{settings.excel_sheet_name}' not found in Excel template file.")
            ws = workbook[settings.excel_sheet_name]
            # Check whether the table name exists
            if settings.excel_table_name not in ws.tables.keys():
                raise ValueError(f"Table '{settings.excel_table_name}' does not exist in sheet "
                                 f"'{settings.excel_sheet_name}'.")
            # Check if Excel template worksheet structure is correct
            tb = ws.tables[settings.excel_table_name]
            from_x, from_y, to_x, to_y = range_boundaries(tb.ref)
            for i in range(len(ReportCreator.COLUMN_NAMES)):
                # TODO only works for 26 columns
                letter = get_column_letter(from_x + i)
                sheet_value = ws[f"{letter}{from_y}"].value
                expected_value = ReportCreator.COLUMN_NAMES[i]
                if sheet_value != ReportCreator.COLUMN_NAMES[i]:
                    raise ValueError(
                        f"Expected '{expected_value}' but '{sheet_value}' found column {letter}1."
                    )
